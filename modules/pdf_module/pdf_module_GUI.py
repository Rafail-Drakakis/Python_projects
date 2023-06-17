import tkinter
import pandas
import PyPDF2
import pdf2docx
import pdf2image
import tabula
import os
import tempfile
import openpyxl
from tkinter import filedialog, simpledialog, messagebox

def merge_pdf_files():
    files = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    if files:
        output_filename = filedialog.asksaveasfilename(filetypes=[("PDF Files", "*.pdf")])
        if output_filename:
            merger = PyPDF2.PdfMerger()
            for file in files:
                merger.append(file)
            merger.write(output_filename)
            merger.close()
        messagebox.showinfo("Conversion complete")

def split_pdf():
    filename = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    if filename:
        while True:
            page_ranges_input = page_ranges_dialog()
            if page_ranges_input:
                page_ranges = [range.strip() for range in page_ranges_input.split(',')]
                try:
                    split_pdf_pages(filename, page_ranges)
                    messagebox.showinfo("Conversion complete")
                    break  # Exit the loop if the conversion is successful
                except ValueError as e:
                    messagebox.showerror("Error", str(e))
                    retry = messagebox.askretrycancel("Retry", "Do you want to select pages again?")
                    if not retry:
                        break  # Exit the loop if the user chooses not to retry
            else:
                break  # Exit the loop if the user cancels page range selection

def split_pdf_pages(filename, page_ranges):
    all_pages = []
    for page_range in page_ranges:
        if '-' in page_range:
            start, end = map(int, page_range.split('-'))
            all_pages.extend(range(start, end + 1))
        else:
            page = int(page_range)
            all_pages.append(page)

    with open(filename, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        num_pages = len(reader.pages)
        if not all(1 <= p <= num_pages for p in all_pages):
            raise ValueError('Invalid page range')
        writer = PyPDF2.PdfWriter()
        for p in all_pages:
            writer.add_page(reader.pages[p - 1])
        output_filename = filedialog.asksaveasfilename(filetypes=[("PDF Files", "*.pdf")])
        if output_filename:
            with open(output_filename, 'wb') as new_file:
                writer.write(new_file)
            messagebox.showinfo("Conversion complete")

def convert_pdf_to_word():
    filenames = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    if filenames:
        for filename in filenames:
            docx_path = filename.replace(".pdf", ".docx")
            pdf2docx.parse(filename, docx_path)
        messagebox.showinfo("Conversion complete")

def convert_pdf_to_images():
    filenames = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    if filenames:
        output_directory = filedialog.askdirectory()  # Ask for output directory
        for filename in filenames:
            images = pdf2image.convert_from_path(filename, dpi=1000)
            for idx, img in enumerate(images):
                image_path = f"{output_directory}/page_{idx + 1}.jpg"
                img.save(image_path, 'JPEG', quality=80)
        messagebox.showinfo("Conversion complete")

def page_ranges_dialog():
    page_ranges = simpledialog.askstring("Page Ranges", "Enter page ranges (e.g., 1, 3-5):")
    return page_ranges

def convert_pdf_to_excel(input_pdf_path, output_excel_path):
    # Convert PDF to Excel
    pdf_data_frames = tabula.read_pdf(input_pdf_path, pages='all')
    with pandas.ExcelWriter(output_excel_path, engine='openpyxl') as excel_writer:
        for i, df in enumerate(pdf_data_frames):
            df.to_excel(excel_writer, sheet_name=f"Sheet{i+1}", index=False)

def merge_excel_sheets(input_excel_path, output_excel_path):
    # Merge Excel sheets
    dfs = pandas.read_excel(input_excel_path, sheet_name=None)
    merged_data_frame = pandas.concat(dfs.values(), ignore_index=True)
    merged_data_frame.to_excel(output_excel_path, index=False)

def shift_empty_cells(output_excel_path):
    # Load output Excel file
    excel_workbook = openpyxl.load_workbook(output_excel_path)
    for worksheet_name in excel_workbook.sheetnames:
        excel_worksheet = excel_workbook[worksheet_name]
        for row in excel_worksheet.iter_rows():
            empty_cell_list = []
            for current_cell in row:
                if current_cell.value is None:
                    empty_cell_list.append(current_cell)
                else:
                    for empty_cell_to_fill in empty_cell_list:
                        empty_cell_to_fill.value = current_cell.value
                        current_cell.value = None
                        empty_cell_list.remove(empty_cell_to_fill)
                        empty_cell_list.append(current_cell)
                        break
    excel_workbook.save(output_excel_path)

def pdf_to_excel():
    # Prompt user to select PDF file
    input_pdf_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    if not input_pdf_path:
        return

    # Ask user for desired output filename and define output Excel file path
    output_excel_path = filedialog.asksaveasfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not output_excel_path:
        return

    # Create temporary Excel file paths
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_excel_path = temp_file.name
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as merged_file:
        merged_excel_path = merged_file.name

    try:
        convert_pdf_to_excel(input_pdf_path, temp_excel_path)
        merge_excel_sheets(temp_excel_path, merged_excel_path)
        os.replace(merged_excel_path, output_excel_path)

        # Ask if user wants to merge empty cells
        merge_option = messagebox.askquestion("Merge Empty Cells", "Do you want to merge empty cells?")
        if merge_option == "yes":
            shift_empty_cells(output_excel_path)

        messagebox.showinfo("Conversion Complete", f"File {output_excel_path} has been created successfully.")
    finally:
        # Remove temporary files
        os.remove(temp_excel_path)
        if os.path.exists(merged_excel_path):
            os.remove(merged_excel_path)

# Application window
app = tkinter.Tk()
app.geometry("520x300")
app.title("PDF Operator")

# Frame
buttons_frame = tkinter.Frame(app)
buttons_frame.pack(padx=10, pady=10)

#Buttons
merge_pdf_button = tkinter.Button(buttons_frame, text="Merge PDF", command=merge_pdf_files)
merge_pdf_button.pack(padx=10, pady=5)

split_pdf_button = tkinter.Button(buttons_frame, text="Split PDF", command=split_pdf)
split_pdf_button.pack(padx=10, pady=5)

convert_pdf_to_images_button = tkinter.Button(buttons_frame, text="Convert PDF to Images", command=convert_pdf_to_images)
convert_pdf_to_images_button.pack(padx=10, pady=5)

convert_pdf_to_word_button = tkinter.Button(buttons_frame, text="Convert PDF to Word", command=convert_pdf_to_word)
convert_pdf_to_word_button.pack(padx=10, pady=5)

pdf_to_excel_button = tkinter.Button(buttons_frame, text="Convert PDF to Excel", command=pdf_to_excel)
pdf_to_excel_button.pack(padx=10, pady=5)

#main loop (start)
app.mainloop()