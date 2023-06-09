#pdf_to_excel.py

import os
import tempfile
import pandas as pd
import tabula
import openpyxl

def merge_excel_sheets(input_excel_path, output_excel_path):
    # Merge Excel sheets
    try:
        dfs = pd.read_excel(input_excel_path, sheet_name=None)
        merged_data_frame = pd.concat(dfs.values(), ignore_index=True)
        merged_data_frame.to_excel(output_excel_path, index=False)
    except FileNotFoundError as e:
        print(f"Error: File not found. {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

def shift_empty_cells(output_excel_path):
    # Load output Excel file
    try:
        excel_workbook = openpyxl.load_workbook(output_excel_path)
        for worksheet_name in excel_workbook.sheetnames:
            excel_worksheet = excel_workbook[worksheet_name]
            for row in excel_worksheet.iter_rows():
                empty_cell_list = []
                for current_cell in row:
                    if current_cell.value is None:
                        empty_cell_list.append(current_cell)
                    else:
                        for empty_cell_to_fill in empty_cell_list:
                            empty_cell_to_fill.value = current_cell.value
                            current_cell.value = None
                            empty_cell_list.remove(empty_cell_to_fill)
                            empty_cell_list.append(current_cell)
                            break
        excel_workbook.save(output_excel_path)
    except FileNotFoundError as e:
        print(f"Error: File not found. {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

def pdf_to_excel(input_pdf_path):
    # Define output Excel file path
    output_excel_path = os.path.splitext(input_pdf_path)[0] + ".xlsx"

    # Create temporary Excel file paths
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_excel_path = temp_file.name
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as merged_file:
        merged_excel_path = merged_file.name

    try:
        # Convert PDF to Excel
        pdf_data_frames = tabula.read_pdf(input_pdf_path, pages='all')
        with pd.ExcelWriter(temp_excel_path, engine='openpyxl') as excel_writer:
            for i, df in enumerate(pdf_data_frames):
                df.to_excel(excel_writer, sheet_name=f"Sheet{i + 1}", index=False)
    
        merge_excel_sheets(temp_excel_path, merged_excel_path)
        os.replace(merged_excel_path, output_excel_path)

        # Check if user wants to shift empty cells
        option = input("Do you want to shift empty cells? (yes/no) ")
        if option == "yes":
            shift_empty_cells(output_excel_path)
    except FileNotFoundError as e:
        print(f"Error: File not found. {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Remove temporary files
        os.remove(temp_excel_path)
        if os.path.exists(merged_excel_path):
            os.remove(merged_excel_path)